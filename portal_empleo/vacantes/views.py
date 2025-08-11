from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from .models import Vacante, Ciudad, Departamento, RegistroCandidato, EstadoAplicacion, ComentarioCandidato
from .forms import VacanteForm, RegistroCandidatoForm, ComentarioCandidatoForm #Añadir ComentarioCandidatoForm
from django.contrib.auth.decorators import login_required # Solicitar ligin para ejecutar funcion
from django.contrib.auth.decorators import permission_required # Solicita login para permisos especificos @permission_required('app_name.add_vacante')
import openpyxl
from django.db.models.functions import Lower
from django.db.models import Count, Q
import unicodedata
from django.db import migrations
from django.views.generic import TemplateView 
from datetime import datetime
import json
import csv
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponseRedirect, QueryDict 
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User 
from django.utils.timezone import localtime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger # Asegúrate que esta línea exista
import shutil
from django.conf import settings
import os






def normalizar_texto(texto):
    """
    Elimina los acentos y normaliza el texto a minúsculas.
    Si el texto es None, devuelve una cadena vacía.
    """
    if texto is None:
        return ""  # Devolver cadena vacía si el texto es None

    return ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    ).lower()
    

def lista_vacantes(request):
    
    # Procesar los parámetros de filtro multiselect
    filtros = {
        'cargo': request.GET.get('cargo', ''),
        'codigo_vacante': request.GET.get('codigo_vacante', ''),
        'rango_salarial_min': request.GET.get('rango_salarial_min', ''),
        'rango_salarial_max': request.GET.get('rango_salarial_max', ''),

        # Campos multiselect (pueden tener múltiples valores)
        'usuario_publicador_list': request.GET.getlist('usuario_publicador'),
        'estado_list': request.GET.getlist('estado'), # Recibe ['True'], ['False'], o ['True', 'False']
        'area_list': request.GET.getlist('area'),
        'modalidad_trabajo_list': request.GET.getlist('modalidad_trabajo'),
        'tipo_contrato_list': request.GET.getlist('tipo_contrato'),
        'jornada_trabajo_list': request.GET.getlist('jornada_trabajo'),
        'tiempo_experiencia_list': request.GET.getlist('tiempo_experiencia'),
        'nivel_estudios_list': request.GET.getlist('nivel_estudios'),
        'departamento_list': request.GET.getlist('departamento'),
        'ciudad_list': request.GET.getlist('ciudad'),
    }

    # Obtener todas las vacantes activas si el usuario no está autenticado
    vacantes = Vacante.objects.filter(estado=True) if not request.user.is_authenticated else Vacante.objects.all()

    usuarios_publicadores = User.objects.all().order_by('username')


    # Filtro para Codigo Vacante (ID Cargo)
    if filtros['codigo_vacante']:
        vacantes = vacantes.filter(codigo_vacante__icontains=filtros['codigo_vacante'].strip())
    
    
    # Filtro para Cargo
    if filtros['cargo']:
        vacantes = vacantes.filter(cargo__icontains=filtros['cargo'].strip())

    # Aplicar filtros multiselect
    if filtros['area_list']:
        vacantes = vacantes.filter(area__in=filtros['area_list'])
        
    # filtro estado    
    if filtros['estado_list']:
        # Convertir los strings 'True'/'False' de la URL a booleanos
        estados_booleanos = []
        if 'True' in filtros['estado_list']:
            estados_booleanos.append(True)
        if 'False' in filtros['estado_list']:
            estados_booleanos.append(False)
            
        if estados_booleanos: # Solo filtra si se seleccionó algo válido
            vacantes = vacantes.filter(estado__in=estados_booleanos)
            
    #filtro usuario publicador
    if filtros['usuario_publicador_list']:
        vacantes = vacantes.filter(usuario_publicador__id__in=filtros['usuario_publicador_list'])

    
    
    if filtros['modalidad_trabajo_list']:
        vacantes = vacantes.filter(modalidad_trabajo__in=filtros['modalidad_trabajo_list'])
    
    if filtros['tipo_contrato_list']:
        vacantes = vacantes.filter(tipo_contrato__in=filtros['tipo_contrato_list'])
    
    if filtros['jornada_trabajo_list']:
        vacantes = vacantes.filter(jornada_trabajo__in=filtros['jornada_trabajo_list'])
    
    if filtros['tiempo_experiencia_list']:
        vacantes = vacantes.filter(tiempo_experiencia__in=filtros['tiempo_experiencia_list'])
    
    if filtros['nivel_estudios_list']:
        vacantes = vacantes.filter(nivel_estudios__in=filtros['nivel_estudios_list'])
    
    if filtros['departamento_list']:
        vacantes = vacantes.filter(departamento__in=filtros['departamento_list'])
    
    # Filtro de Ciudad (simplificado, solo IDs numéricos)
    if filtros['ciudad_list']:
        from django.db.models import Q
        # Directamente filtrar por los IDs numéricos recibidos
        filtro_ciudades = Q(ciudad__in=filtros['ciudad_list'])
        
        # Aplicar el filtro combinado si hay alguna condición
        if filtro_ciudades:
            vacantes = vacantes.filter(filtro_ciudades)

    # Filtro para Rango Salarial
    rango_salarial_min = filtros['rango_salarial_min']
    rango_salarial_max = filtros['rango_salarial_max']
    
    if rango_salarial_min and rango_salarial_max:
        vacantes = vacantes.filter(
            rango_salarial__gte=rango_salarial_min,
            rango_salarial__lte=rango_salarial_max
        )
    elif rango_salarial_min:
        vacantes = vacantes.filter(rango_salarial__gte=rango_salarial_min)
    elif rango_salarial_max:
        vacantes = vacantes.filter(rango_salarial__lte=rango_salarial_max)

    # Agregar el conteo de candidatos aplicados
    vacantes = vacantes.annotate(num_candidatos=Count('candidatos'))

    # Mantener el orden alfabético por cargo
    vacantes = sorted(vacantes, key=lambda v: v.cargo.lower())

    # --- INICIO: Lógica de Paginación ---

    # Obtener ítems por página (con un valor por defecto)
    try:
        items_por_pagina = int(request.GET.get('por_pagina', 10)) # Usa el valor del form, default 10
    except ValueError:
        items_por_pagina = 10 # Default si el valor no es un número válido

    # Instanciar el Paginator con la lista ya filtrada y ordenada
    paginator = Paginator(vacantes, items_por_pagina)

    # Obtener el número de página solicitado
    page_number = request.GET.get('page')

    # Obtener el objeto Page (page_obj) y manejar errores
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        # Si page no es un entero, entrega la primera página.
        page_obj = paginator.page(1)
    except EmptyPage:
        # Si page está fuera de rango (p.ej. 9999), entrega la última página de resultados.
        page_obj = paginator.page(paginator.num_pages)

    # --- FIN: Lógica de Paginación ---

    # Obtener opciones de los campos con choices
    form = VacanteForm()
    choices_context = {
        'area_choices': form.fields['area'].choices,
        'modalidad_trabajo_choices': form.fields['modalidad_trabajo'].choices,
        'tipo_contrato_choices': form.fields['tipo_contrato'].choices,
        'jornada_trabajo_choices': form.fields['jornada_trabajo'].choices,
        'tiempo_experiencia_choices': form.fields['tiempo_experiencia'].choices,
        'nivel_estudios_choices': form.fields['nivel_estudios'].choices,
        'departamento_choices': form.fields['departamento'].choices,
    }
    
    # Obtener el departamento "Cundinamarca" y sus ciudades
    cundinamarca = Departamento.objects.filter(nombre="Cundinamarca").first()
    ciudades_cundinamarca = []
    if cundinamarca:
        # Excluimos Bogotá de la lista de Cundinamarca para evitar duplicados en el select
        ciudades_cundinamarca = Ciudad.objects.filter(departamento=cundinamarca).exclude(id=12117)

    # Preparar la información de ciudades seleccionadas para mostrar en filtros activos
    # (simplificado, basado directamente en filtros['ciudad_list'])
    filtros['ciudad_info'] = []
    if filtros['ciudad_list']:
        # Obtener los objetos Ciudad correspondientes a los IDs seleccionados
        ciudades_seleccionadas = Ciudad.objects.filter(id__in=filtros['ciudad_list'])
        for ciudad in ciudades_seleccionadas:
            filtros['ciudad_info'].append({ # Usar siempre el ID numérico
                'id': str(ciudad.id),
                'nombre': ciudad.nombre
            })

    # Renderizar la plantilla con el objeto de página y otros datos de contexto
    return render(request, "vacantes/lista.html", {
        "page_obj": page_obj, # <--- ¡Clave! Pasar el objeto de página
        "items_por_pagina": items_por_pagina, # Para el selector <select>
        "filtros": filtros,
        "ciudades_cundinamarca": ciudades_cundinamarca,
        "usuarios_publicadores": usuarios_publicadores,
        **choices_context  # Pasar los choices al template
    })# API para obtener ciudades por departamentos (añadir a urls.py)
def ciudades_por_departamentos(request):
    departamentos = request.GET.get('departamentos', '').split(',')
    ciudades = Ciudad.objects.filter(departamento__in=departamentos).values('id', 'nombre')
    return JsonResponse(list(ciudades), safe=False)# para cambiar el estado de la vacante
def cambiar_estado_vacante(request, vacante_id):
    if request.method == 'POST':
        vacante = get_object_or_404(Vacante, id=vacante_id)
        vacante.estado = not vacante.estado
        # Aquí podrías querer guardar quién actualizó, si tienes esa lógica
        # vacante.usuario_actualizador = request.user 
        vacante.save()

        # Reconstruir los parámetros de la query string a partir del POST
        query_params = QueryDict(mutable=True)
        for key, value_list in request.POST.lists():
            # Excluir el token CSRF y cualquier otro campo específico del POST
            # que no sea un filtro.
            if key not in ['csrfmiddlewaretoken']: # Añade aquí otros campos a excluir si es necesario
                for value in value_list:
                    query_params.appendlist(key, value)
        
        # Construir la URL de redirección con los filtros
        redirect_url = reverse('lista_vacantes') # Asegúrate que 'lista_vacantes' es el name de tu URL
        if query_params:
            redirect_url += '?' + query_params.urlencode()
            
        return redirect(redirect_url)
    
    # Si no es POST, simplemente redirigir a la lista (o manejar como un error)
    return redirect(reverse('lista_vacantes'))


# Create your views here.



# Archivo de migración: 0005_clean_numero_puestos.py



def limpiar_numero_puestos(apps, schema_editor):
    Vacante = apps.get_model('proyecto_negocio', 'Vacante')
    # Convertir los valores de 'n/a' a 0
    Vacante.objects.filter(numero_puestos='n/a').update(numero_puestos=0)
    # Convertir los valores no numéricos en 0
    for vacante in Vacante.objects.all():
        try:
            vacante.numero_puestos = int(vacante.numero_puestos)
            vacante.save()
        except ValueError:
            vacante.numero_puestos = 0
            vacante.save()

class Migration(migrations.Migration):

    dependencies = [
        ('proyecto_negocio', '0004_remove_vacante_empresa_remove_vacante_estudios_and_more'),
    ]

    operations = [
        migrations.RunPython(limpiar_numero_puestos),
    ]


#Agregado por Jose
@login_required
def eliminar_vacante(request, id):
    """Elimina la vacante con el ID proporcionado."""
    if request.method == 'POST':
        vacante = get_object_or_404(Vacante, id=id)
        vacante.delete()
        messages.success(request, 'La vacante ha sido eliminada correctamente.')
    # Cambia 'nombre_de_tu_vista_principal' por la vista principal, por ejemplo, 'lista_vacantes'
    return redirect('lista_vacantes')

@login_required
def agregar_vacante(request):
    if request.method == 'POST':
        form = VacanteForm(request.POST, user=request.user)  # Pasamos el usuario
        if form.is_valid():
            vacante = form.save(commit=False)  # No guardamos aún en la BD
            vacante.usuario_publicador = request.user  # Asignamos el usuario
            vacante.save()  # Ahora sí guardamos
            return redirect('lista_vacantes')  # Redirige a la lista de vacantes
    else:
        form = VacanteForm(user=request.user)  # Pasamos el usuario también en GET

    return render(request, 'vacantes/agregar_vacante.html', {'form': form})

def inicio(request):
    return render(request, "paginas/inicio.html")

@login_required
def editar_vacante(request, id):
    vacante = get_object_or_404(Vacante, id=id)
    
    if request.method == 'POST':
        form = VacanteForm(request.POST, instance=vacante)
        if form.is_valid():
            vacante = form.save(commit=False)  # No guardamos aún
            vacante.usuario_actualizador = request.user  # Guardamos quién actualiza
            vacante.save()  # Guardamos la vacante con el usuario actualizador
            return redirect('lista_vacantes')

    else:
        form = VacanteForm(instance=vacante)
    
    return render(request, 'vacantes/editar_vacante.html', {'form': form, 'vacante': vacante})


#Registro de candidatos (modificado)

from django.urls import reverse

def registro_candidato_view(request):
    selected_vacantes = request.GET.get('selected_vacantes', '').split(',')
    if request.method == "POST":
        form = RegistroCandidatoForm(request.POST)
        if form.is_valid():
            messages.success(request, '¡Tu registro ha sido completado exitosamente! Por favor pasa con uno de nuestros profesionales para completar la postulación a las vacantes.')
            form.save()
            return redirect(reverse('registration_guide'))  # Redirigir a la Guía de Registro SISE
    else:
        form = RegistroCandidatoForm()

    return render(request, 'registro_candidato.html', {'form': form, 'selected_vacantes': selected_vacantes})



# def registro_candidato_view(request):
#     selected_vacantes_ids = request.GET.get('selected_vacantes', '').split(',')
#     selected_vacantes = Vacante.objects.filter(id__in=selected_vacantes_ids)

#     if request.method == "POST":
#         form = RegistroCandidatoForm(request.POST)
#         if form.is_valid():
#             form.save()
#             return redirect('inicio')  # Redirigir a una página de éxito
#     else:
#         form = RegistroCandidatoForm()

#     return render(request, 'registro_candidato.html', {
#         'form': form,
#         'selected_vacantes': selected_vacantes
#     })





#Descargar Excel

def descargar_excel(request):
    """Genera y descarga un archivo Excel con todos los datos de las vacantes, incluyendo descripción completa."""
    import openpyxl
    from django.http import HttpResponse
    from django.utils.timezone import now
    from django.db import connection
    import logging

    # Configurar logging
    logger = logging.getLogger('django')
    logger.info("Iniciando descarga de Excel con descripción completa...")

    # Crear el libro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vacantes"
    
    # Encabezados estándar
    headers = [
        "Código Vacante", "Cargo", "Área", "Número de Puestos", 
        "Modalidad Trabajo", "Tipo Contrato", "Jornada Trabajo",
        "Descripción", "Experiencia", "Nivel Estudios",
        "Departamento", "Ciudad", "Rango Salarial",
        "Empresa", "Estado", "Fecha Publicación",
        "Tipo Vacante", "Citacion Convocatoria", "Cantidad Candidatos Registrados",
        "Usuario Publicador", "Fecha Creación", "Fecha Actualización"
    ]
    ws.append(headers)
    
    # Consulta de vacantes
    vacantes = Vacante.objects.annotate(cantidad_candidatos=Count('candidatos'))
    
    for vacante in vacantes:
        # Añadir fila al Excel, con descripción completa
        ws.append([
            vacante.codigo_vacante,
            vacante.cargo,
            vacante.area if vacante.area else "No especificado",
            vacante.numero_puestos if vacante.numero_puestos else "No especificado",
            vacante.modalidad_trabajo,
            vacante.tipo_contrato,
            vacante.jornada_trabajo,
            # Mostrar la descripción completa sin truncar
            vacante.descripcion_vacante,
            vacante.tiempo_experiencia,
            vacante.nivel_estudios,
            vacante.departamento.nombre if vacante.departamento else "No especificado",
            vacante.ciudad.nombre if vacante.ciudad else "No especificado",
            vacante.rango_salarial if vacante.rango_salarial else "No especificado",
            vacante.empresa_usuaria,
            "Activa" if vacante.estado else "Inactiva",
            vacante.fecha_publicacion.strftime("%d/%m/%Y"),
            vacante.tipo_vacante,
            vacante.citacion_convocatoria if vacante.citacion_convocatoria else "No aplica",
            vacante.cantidad_candidatos,
            vacante.usuario_publicador.username if vacante.usuario_publicador else "Sin registro",
            vacante.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
            vacante.fecha_actualizacion.strftime("%d/%m/%Y %H:%M") if vacante.fecha_actualizacion else "Sin actualización"
        ])
    
    # Ajustar el ancho de las columnas, pero permitiendo que la descripción sea ancha
    for column_index, column in enumerate(ws.columns):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        
        # Configuración especial para la columna de descripción (columna H, índice 7)
        if column_index == 7:  # La columna de descripción
            ws.column_dimensions[column_letter].width = 100  # Ancho fijo amplio para descripción
        else:
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Configurar el estilo para permitir texto multilínea y ajuste de texto
    from openpyxl.styles import Alignment
    for row in ws.iter_rows(min_row=2):  # Comenzar desde la fila 2 (después de los encabezados)
        # Aplicar alineación y ajuste de texto a la celda de descripción (columna H, índice 7)
        row[7].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Ajustar la altura de las filas para mostrar mejor el texto multilínea
    for row_index in range(2, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 90  # Altura fija para todas las filas con datos
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="vacantes_completo.xlsx"'
    wb.save(response)
    return response

# def lista_candidatos(request, id):
#     # Obtener la vacante específica
#     vacante = get_object_or_404(Vacante, id=id)
    
#     # Obtener la lista de candidatos relacionados con esta vacante
#     candidatos = vacante.candidatos.all()

#     # Traer o crear el estado de cada aplicación
#     for candidato in candidatos:
#         estado_aplicacion, creado = EstadoAplicacion.objects.get_or_create(
#             candidato=candidato, 
#             vacante=vacante,
#             defaults={'estado': 'No visto'}
#         )
#         candidato.estado_aplicacion = estado_aplicacion.estado
    
#     # Renderizar la información en el template
#     return render(request, 'vacantes/lista_candidatos.html', {'vacante': vacante, 'candidatos': candidatos})












# views.py



def lista_candidatos(request, id):
    vacante = get_object_or_404(Vacante, id=id)
    candidatos = vacante.candidatos.all()

    # Obtener todos los filtros de la solicitud GET
    filtros = {}
    for key, value in request.GET.items():
        if value:  # Solo si el valor no está vacío
            filtros[key] = value

    print(f"Filtros recibidos: {filtros}")  # debug
    
        # Obtener todos los comentarios para cada candidato (sin filtrar)
    todos_los_comentarios = {}
    for candidato in candidatos:
        todos_los_comentarios[candidato.id] = candidato.comentarios.all()

    # Aplicar los filtros a los candidatos
    if filtros:
        q_objects = Q()
        for key, value in filtros.items():
            if key == 'departamento':
                q_objects &= Q(departamento__id=value)  # AND operator for relational fields
            elif key == 'ciudad':
                q_objects &= Q(ciudad__id=value)  # AND operator for relational fields
            elif key == 'filterNombre':
                palabras = value.split()
                for palabra in palabras:
                    q_objects &= (Q(nombres__icontains=palabra) | Q(apellidos__icontains=palabra))
            elif key == 'filterTipoDocumento':
                q_objects &= Q(tipo_documento=value)  # AND operator
            elif key == 'filterNumeroDocumento':
                q_objects &= Q(numero_documento__icontains=value)  # AND operator
            elif key == 'filterSexo':
                q_objects &= Q(sexo=value)  # AND operator
            elif key == 'filterNacimientoDesde':
                try:
                    value = datetime.strptime(value, '%Y-%m-%d').date()  # Convert to date
                    q_objects &= Q(fecha_nacimiento__gte=value)  # AND operator
                except ValueError:
                    pass  # Ignorar el filtro si el formato es incorrecto
            elif key == 'filterNacimientoHasta':
                try:
                    value = datetime.strptime(value, '%Y-%m-%d').date()  # Convert to date
                    q_objects &= Q(fecha_nacimiento__lte=value)  # AND operator
                except ValueError:
                    pass  # Ignorar el filtro si el formato es incorrecto
            elif key == 'filterEstudioMinimo':
                q_objects &= Q(formacion_academica=value)  # AND operator
            elif key == 'filterPrograma':
                q_objects &= Q(programa_academico__icontains=value)  # AND operator
            elif key == 'filterHorario':
                q_objects &= Q(horario_interesado=value)  # AND operator
            elif key == 'filterSalarioMin':
                try:
                    value = int(value)
                    q_objects &= Q(aspiracion_salarial__gte=value)  # AND operator
                except ValueError:
                    pass  # Ignorar el filtro si el formato es incorrecto
            elif key == 'filterSalarioMax':
                try:
                    value = int(value)
                    q_objects &= Q(aspiracion_salarial__lte=value)  # AND operator
                except ValueError:
                    pass  # Ignorar el filtro si el formato es incorrecto
            elif key == 'filterNombreFeria':
                q_objects &= Q(feria__icontains=value)  # AND operator
            elif key == 'filterFeriaDesde':
                try:
                    value = datetime.strptime(value, '%Y-%m-%d').date()  # Convert to date
                    q_objects &= Q(fecha_feria__gte=value)  # AND operator
                except ValueError:
                    pass  # Ignorar el filtro si el formato es incorrecto
            elif key == 'filterFeriaHasta':
                try:
                    value = datetime.strptime(value, '%Y-%m-%d').date()  # Convert to date
                    q_objects &= Q(fecha_feria__lte=value)  # AND operator
                except ValueError:
                    pass  # Ignorar el filtro si el formato es incorrecto
            elif key == 'filterTecnico':
                q_objects &= Q(tecnico_seleccion=value)  # AND operator
            elif key == 'filterSISE':
                q_objects &= Q(registrado_en_sise=value)  # AND operator

        candidatos = candidatos.filter(q_objects)
        print(f"Candidatos después de aplicar todos los filtros: {candidatos.count()}")  # debug

    # Obtener o crear el estado de cada aplicación
    for candidato in candidatos:
        estado_aplicacion, _ = EstadoAplicacion.objects.get_or_create(
            candidato=candidato,
            vacante=vacante,
            defaults={'estado': 'No visto'}
        )
        candidato.estado_aplicacion = estado_aplicacion.estado

    # preparar la data de los filtros para la plantilla
    departamento_choices = Departamento.objects.all().values_list('id', 'nombre')
    ciudad_choices = []
    if filtros.get('departamento'):
        ciudad_choices = Ciudad.objects.filter(departamento_id=filtros['departamento']).values_list('id', 'nombre')

    # se crea un nuevo diccionario para la informacion de la plantilla
    filtros_para_la_plantilla = {
        key: value for key, value in filtros.items() if key not in ['departamento', 'ciudad']
    }
    filtros_para_la_plantilla['departamento'] = int(filtros['departamento']) if filtros.get('departamento') else ''
    filtros_para_la_plantilla['ciudad'] = int(filtros['ciudad']) if filtros.get('ciudad') else ''

    if filtros_para_la_plantilla.get('ciudad'):
        filtros_para_la_plantilla['ciudad_nombre'] = Ciudad.objects.get(id=filtros_para_la_plantilla['ciudad']).nombre
    else:
        filtros_para_la_plantilla['ciudad_nombre'] = ''

    contexto = {
        'vacante': vacante,
        'candidatos': candidatos,
        'departamento_choices': departamento_choices,
        'ciudad_choices': ciudad_choices,
        'filtros': filtros_para_la_plantilla,
        'todos_los_comentarios': todos_los_comentarios,  # Pasar todos los comentarios al template
        'form_comentario': ComentarioCandidatoForm() # Se crea una instancia del formulario
    }

    return render(request, 'vacantes/lista_candidatos.html', contexto)





def lista_registros(request):
    query = request.GET.get('q', '')
    order_by = request.GET.get('order_by', 'fecha_feria')
    order_dir = request.GET.get('order_dir', 'asc')

    registros = RegistroCandidato.objects.all()

    # Definir campos válidos para ordenación
    valid_order_fields = {
        'nombres': 'nombres_lower',
        'apellidos': 'apellidos_lower',
        'numero_documento': 'documento_lower',
        'feria': 'feria',
        'fecha_feria': 'fecha_feria'
    }

    # Si la ordenación requiere anotaciones, añadirlas
    if order_by in ['nombres', 'apellidos', 'numero_documento'] or query:
        registros = registros.annotate(
            nombres_lower=Lower('nombres'),
            apellidos_lower=Lower('apellidos'),
            documento_lower=Lower('numero_documento')
        )

    # Si hay búsqueda, filtrar
    if query:
        registros = registros.filter(
            Q(nombres_lower__icontains=query) |
            Q(apellidos_lower__icontains=query) |
            Q(documento_lower__icontains=query)
        )

    # Determinar campo de orden válido
    sort_field = valid_order_fields.get(order_by, 'fecha_feria')

    # Aplicar la ordenación
    if order_dir == 'desc':
        registros = registros.order_by(f'-{sort_field}')
    else:
        registros = registros.order_by(sort_field)
        
    # --- INICIO: Lógica de Paginación ---
    try:
        # Leer 'por_pagina' de GET, default a 10 si no está o es inválido (igual que en lista_vacantes)
        items_por_pagina = int(request.GET.get('por_pagina', 10)) 
    except ValueError:
        items_por_pagina = 10 # Default si el valor no es un número válido (igual que en lista_vacantes)

    paginator = Paginator(registros, items_por_pagina)
    page_number = request.GET.get('page')

    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        # Si page no es un entero, entrega la primera página.
        page_obj = paginator.page(1)
    except EmptyPage:
        # Si page está fuera de rango, entrega la última página de resultados.
        page_obj = paginator.page(paginator.num_pages)
    # --- FIN: Lógica de Paginación ---


    context = {
        'page_obj': page_obj, # <--- Pasar el objeto de página en lugar de la lista completa
        'items_por_pagina': items_por_pagina, # Para el selector <select>
        'query': query,
        'order_by': order_by,
        'order_dir': order_dir
    }
    return render(request, 'lista_registros.html', context)

def editar_registro(request, pk):
    registro = get_object_or_404(RegistroCandidato, pk=pk)
    
    if request.method == 'POST':
        form = RegistroCandidatoForm(request.POST, instance=registro)
        if form.is_valid():
            form.save()
            return redirect('lista_registros')
    else:
        form = RegistroCandidatoForm(instance=registro)
        # Formatear las fechas para el input type="date"
        if registro.fecha_nacimiento:
            form.fields['fecha_nacimiento'].initial = registro.fecha_nacimiento.strftime('%Y-%m-%d')
        if registro.fecha_feria:
            form.fields['fecha_feria'].initial = registro.fecha_feria.strftime('%Y-%m-%d')
        
    context = {
        'form': form,
        'selected_vacantes': registro.vacantes_disponibles.all(),
        'is_edit': True
    }
    return render(request, 'registro_candidato.html', context)

# para cargar ciudades por departamento

def cargar_ciudades(request):
    departamento_id = request.GET.get("departamento_id")
    ciudades = Ciudad.objects.filter(departamento_id=departamento_id).values("id", "nombre")
    return JsonResponse(list(ciudades), safe=False)


#def candidatos_por_vacante(request, vacante_id):
#    vacante = get_object_or_404(Vacante, id=vacante_id)
#    candidatos = vacante.candidatos.all()
#    return render(request, 'candidatos_por_vacante.html', {'vacante': vacante, 'candidatos': candidatos})


class RegistrationGuideView(TemplateView):
    template_name = 'registration/guide.html'
    
def exportar_candidatos_excel(request):
    """Genera y descarga un archivo Excel con la lista de candidatos registrados y las vacantes a las que aplicaron."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidatos"

    # Orden actualizado según lo solicitado
    columnas = [
        "Tipo Feria", "Feria", "Fecha Feria", "Sexo", "Tipo Documento", "Número Documento", 
        "Nombres", "Apellidos", "Número Celular", "Correo Electrónico", "Fecha Nacimiento", 
        "Formación Académica", "Semestre/Grado", "Interés Práctica", "Programa Académico",
        "Experiencia Laboral", "Interés Ocupacional", "Departamento o Ciudad", "Localidad o Municipio", 
        "Candidato Discapacidad", "Tipo Discapacidad", "Horario Interesado", 
        "Aspiración Salarial", "Técnico Selección", "Vacantes Aplicadas", 
        "Código SISE", "Empresa SISE", "Cargo SISE"
    ]
    
    ws.append(columnas)

    candidatos = RegistroCandidato.objects.all()

    for candidato in candidatos:
        # Preparamos los datos de vacantes disponibles
        vacantes_disponibles = ""
        
        if hasattr(candidato, 'vacantes_disponibles') and candidato.vacantes_disponibles.exists():
            vacantes_list = [f"{vacante.cargo}" for vacante in candidato.vacantes_disponibles.all()]
            vacantes_disponibles = ", ".join(vacantes_list) if vacantes_list else ""
        
        # Extraer valores seguros para todos los campos
        tipo_feria = getattr(candidato, 'tipo_feria', "") or ""
        feria = getattr(candidato, 'feria', "") or ""
        fecha_feria = ""
        if hasattr(candidato, 'fecha_feria') and candidato.fecha_feria:
            try:
                fecha_feria = candidato.fecha_feria.strftime("%Y-%m-%d")
            except:
                pass
        
        try:
            sexo_display = candidato.get_sexo_display() or ""
        except (AttributeError, ValueError):
            sexo_display = getattr(candidato, 'sexo', "") or ""
            
        try:
            tipo_documento_display = candidato.get_tipo_documento_display() or ""
        except (AttributeError, ValueError):
            tipo_documento_display = getattr(candidato, 'tipo_documento', "") or ""
        
        numero_documento = getattr(candidato, 'numero_documento', "") or ""
        nombres = getattr(candidato, 'nombres', "") or ""
        apellidos = getattr(candidato, 'apellidos', "") or ""
        numero_celular = getattr(candidato, 'numero_celular', "") or ""
        correo_electronico = getattr(candidato, 'correo_electronico', "") or ""
        
        fecha_nacimiento = ""
        if hasattr(candidato, 'fecha_nacimiento') and candidato.fecha_nacimiento:
            try:
                fecha_nacimiento = candidato.fecha_nacimiento.strftime("%Y-%m-%d")
            except:
                pass
        
        try:
            formacion_academica = candidato.get_formacion_academica_display() or ""
        except (AttributeError, ValueError):
            formacion_academica = getattr(candidato, 'formacion_academica', "") or ""
        
        semestre_grado = getattr(candidato, 'semestre_grado', "") or ""
        
        # Manejar el campo interes_practica según lo solicitado
        interes_practica = ""
        if hasattr(candidato, 'interes_practica') and candidato.interes_practica is True:
            interes_practica = "SI"
        
        programa_academico = getattr(candidato, 'programa_academico', "") or ""
        experiencia_laboral = getattr(candidato, 'experiencia_laboral', "") or ""
        interes_ocupacional = getattr(candidato, 'interes_ocupacional', "") or ""
        
        departamento = ""
        if hasattr(candidato, 'departamento') and candidato.departamento:
            departamento = str(candidato.departamento)
            
        ciudad = ""
        if hasattr(candidato, 'ciudad') and candidato.ciudad:
            ciudad = str(candidato.ciudad)
        
        try:
            candidato_discapacidad = candidato.get_candidato_discapacidad_display() or ""
        except (AttributeError, ValueError):
            candidato_discapacidad = getattr(candidato, 'candidato_discapacidad', "") or ""
            
        tipo_discapacidad = getattr(candidato, 'tipo_discapacidad', "") or ""
        
        try:
            horario_interesado = candidato.get_horario_interesado_display() or ""
        except (AttributeError, ValueError):
            horario_interesado = getattr(candidato, 'horario_interesado', "") or ""
            
        aspiracion_salarial = getattr(candidato, 'aspiracion_salarial', "") or ""
        
        try:
            tecnico_seleccion = candidato.get_tecnico_seleccion_display() or ""
        except (AttributeError, ValueError):
            tecnico_seleccion = getattr(candidato, 'tecnico_seleccion', "") or ""
        
        # Usar los campos codigo_sise, empresa_sise y cargo_sise directamente del modelo
        codigo_sise = getattr(candidato, 'codigo_sise', "") or ""
        empresa_sise = getattr(candidato, 'empresa_sise', "") or ""
        cargo_sise = getattr(candidato, 'cargo_sise', "") or ""
        
        # Añadir todos los campos en el orden especificado
        ws.append([
            tipo_feria,
            feria,
            fecha_feria,
            sexo_display,
            tipo_documento_display,
            numero_documento,
            nombres,
            apellidos,
            numero_celular,
            correo_electronico,
            fecha_nacimiento,
            formacion_academica,
            semestre_grado,
            interes_practica,
            programa_academico,
            experiencia_laboral,
            interes_ocupacional,
            departamento,
            ciudad,
            candidato_discapacidad,
            tipo_discapacidad,
            horario_interesado,
            aspiracion_salarial,
            tecnico_seleccion,
            vacantes_disponibles,
            codigo_sise,
            empresa_sise,
            cargo_sise
        ])
    
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="candidatos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@require_POST
@require_POST
def descargar_candidatos(request):
    # Obtener IDs de candidatos seleccionados
    selected_ids = json.loads(request.POST.get('selected_candidates', '[]'))
    
    if not selected_ids:
        return HttpResponse('No se seleccionaron candidatos', status=400)
    
    # Obtener candidatos seleccionados
    candidatos = RegistroCandidato.objects.filter(id__in=selected_ids)
    
    # Crear nuevo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatos"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    centrado = Alignment(horizontal="center", vertical="center")
    borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir encabezados
    headers = [
        'Nombres', 'Apellidos', 'Tipo Documento', 'Número Documento',
        'Fecha Nacimiento', 'Sexo', 'Celular', 'Correo',
        'departamento', 'ciudad', 'Formación Académica', 'Programa Académico',
        'Experiencia Laboral', 'Interés Ocupacional', 'Horario Interesado',
        'Aspiración Salarial', 'Candidato Discapacidad', 'Tipo Discapacidad',
        'Registrado en SISE', 'Técnico Selección', 'Feria', 'Fecha Feria'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centrado
        cell.border = borde
    
    # Escribir datos
    for row, candidato in enumerate(candidatos, 2):
        datos = [
            candidato.nombres,
            candidato.apellidos,
            candidato.get_tipo_documento_display(),
            candidato.numero_documento,
            candidato.fecha_nacimiento.strftime('%Y-%m-%d') if candidato.fecha_nacimiento else '',
            candidato.get_sexo_display(),
            candidato.numero_celular,
            candidato.correo_electronico,
            str(candidato.departamento),  # ✅ Convertido a string
            str(candidato.ciudad),  # ✅ Convertido a string
            candidato.get_formacion_academica_display(),
            candidato.programa_academico,
            candidato.experiencia_laboral,
            candidato.interes_ocupacional,
            candidato.get_horario_interesado_display(),
            candidato.aspiracion_salarial,
            candidato.get_candidato_discapacidad_display(),
            candidato.tipo_discapacidad,
            candidato.get_registrado_en_sise_display(),
            candidato.get_tecnico_seleccion_display(),
            candidato.feria,
            candidato.fecha_feria.strftime('%Y-%m-%d') if candidato.fecha_feria else ''
        ]
        
        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.alignment = centrado
            cell.border = borde
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Columnas más anchas para campos de texto largo
    ws.column_dimensions['L'].width = 30  # Experiencia Laboral
    ws.column_dimensions['M'].width = 30  # Interés Ocupacional
    
    # Congelar panel superior
    ws.freeze_panes = 'A2'
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="candidatos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'},
    )
    
    # Guardar el archivo
    wb.save(response)
    
    return response

# Función para cambiar el estado manualmente con los botones
def cambiar_estado_aplicacion(request, vacante_id, candidato_id, nuevo_estado):
    vacante = get_object_or_404(Vacante, id=vacante_id)
    candidato = get_object_or_404(RegistroCandidato, id=candidato_id)
    
    # Obtener o crear el estado de la aplicación
    estado_aplicacion, creado = EstadoAplicacion.objects.get_or_create(
        candidato=candidato, 
        vacante=vacante
    )
    estado_aplicacion.estado = nuevo_estado
    estado_aplicacion.save()

    messages.success(request, f"Estado cambiado a {nuevo_estado}.")
    return redirect('lista_candidatos', id=vacante.id)

# Función para cambiar el estado al Ver Hoja de Vida (solo si es "No visto")
@csrf_exempt
def cambiar_estado_a_visto(request, vacante_id, candidato_id):
    if request.method == "POST":
        vacante = get_object_or_404(Vacante, id=vacante_id)
        candidato = get_object_or_404(RegistroCandidato, id=candidato_id)
        
        estado_aplicacion, creado = EstadoAplicacion.objects.get_or_create(
            candidato=candidato, 
            vacante=vacante
        )
        
        # Solo cambiar a "Visto" si el estado actual es "No visto"
        if estado_aplicacion.estado == "No visto":
            estado_aplicacion.estado = "Visto"
            estado_aplicacion.save()
            return JsonResponse({'success': True, 'nuevo_estado': 'Visto'})
        
        return JsonResponse({'success': False, 'message': 'El estado no era "No visto"'})
    return JsonResponse({'error': 'Método no permitido'}, status=405)

# Función para manejar la vista de Hoja de Vida
def ver_hoja_de_vida(request, vacante_id, candidato_id):
    vacante = get_object_or_404(Vacante, id=vacante_id)
    candidato = get_object_or_404(RegistroCandidato, id=candidato_id)
    
    # Obtener o crear el estado de la aplicación
    estado_aplicacion, creado = EstadoAplicacion.objects.get_or_create(
        candidato=candidato, 
        vacante=vacante
    )

    # Si el estado actual es "No visto", cambiar a "Visto"
    if estado_aplicacion.estado == "No visto":
        estado_aplicacion.estado = "Visto"
        estado_aplicacion.save()

    # Redirigir a la misma vista de candidatos o mostrar la hoja de vida
    return HttpResponseRedirect(reverse('lista_candidatos', args=[vacante.id]))

# Función para cambiar el estado a "Visto" solo si es "No visto"
@csrf_exempt
def cambiar_a_visto(request, vacante_id, candidato_id):
    if request.method == "POST":
        vacante = get_object_or_404(Vacante, id=vacante_id)
        candidato = get_object_or_404(RegistroCandidato, id=candidato_id)
        
        # Obtener o crear el estado de la aplicación
        estado_aplicacion, creado = EstadoAplicacion.objects.get_or_create(
            candidato=candidato, 
            vacante=vacante
        )
        
        # Solo cambiar a "Visto" si el estado actual es "No visto"
        if estado_aplicacion.estado == "No visto":
            estado_aplicacion.estado = "Visto"
            estado_aplicacion.save()
            return JsonResponse({'success': True, 'nuevo_estado': 'Visto'})
        
        # Si no se cambia el estado
        return JsonResponse({'success': False, 'message': 'El estado no era "No visto"'})
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)







@login_required
def obtener_comentarios(request, vacante_id, candidato_id):
    """
    Retorna los comentarios de un candidato en formato JSON.
    """
    comentarios = ComentarioCandidato.objects.filter(candidato_id=candidato_id).order_by('-fecha_creacion')
    data = {
        'comentarios': [
            {
                'comentario': comentario.comentario,
                'usuario': comentario.usuario.username if comentario.usuario else 'Desconocido',
                'fecha_creacion': comentario.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S'),
            }
            for comentario in comentarios
        ]
    }
    return JsonResponse(data)

@login_required
def guardar_comentario(request, vacante_id, candidato_id):
    """
    Guarda un nuevo comentario para un candidato específico.
    """
    if request.method == 'POST':
        candidato = get_object_or_404(RegistroCandidato, id=candidato_id)
        comentario_text = request.POST.get('comentario')
        if comentario_text:
            comentario = ComentarioCandidato.objects.create(
                candidato=candidato,
                usuario=request.user,
                comentario=comentario_text
            )
            comentario.save()
            data = {'success': True}
            return JsonResponse(data)
        else:
            data = {'success': False}
            return JsonResponse(data)

    return JsonResponse({'success': False})


def backup_base_datos(request):
    # Carpeta de backups
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)

    # Nombre del archivo con fecha y hora
    fecha_hora = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_backup = f'backup_{fecha_hora}.sqlite3'
    ruta_backup = os.path.join(backup_dir, nombre_backup)

    # Copiar la base de datos al backup
    base_datos = os.path.join(settings.BASE_DIR, 'db.sqlite3')
    shutil.copy2(base_datos, ruta_backup)

    messages.success(request, f'Copia de seguridad creada: {nombre_backup}')
    return redirect('backup_estado')


def backup_estado(request):
    # Revisar backups existentes
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    archivos = os.listdir(backup_dir) if os.path.exists(backup_dir) else []

    # Ordenar por fecha descendente
    archivos.sort(reverse=True)

    # Obtener el último backup
    ultimo_backup = archivos[0] if archivos else 'No hay copias de seguridad aún.'

    return render(request, 'paginas/backup_estado.html', {'ultimo_backup': ultimo_backup})
